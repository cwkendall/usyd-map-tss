"""One-time (idempotent) helper: add/refresh a 'Buildings' sheet in the workbook.

Keyed by Building Code (as it appears in 'Map Data', e.g. 'A10ma') or, for
code-less off-campus sites, by the exact 'Location / Building' name. The build
pipeline derives buildings.json from this sheet and joins it to facility rows.

Coordinates were seeded from OpenStreetMap (Nominatim/Overpass). Rows flagged
VERIFY in Notes are estimates the data owner should confirm. Edit lat/lon here
to move pins; everything downstream regenerates from this sheet.

Run sandboxed:  uv run --with openpyxl python scripts/add-buildings-sheet.py
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

WB = os.path.join(os.path.dirname(__file__), "..", "TSS-CRF-MapData.xlsx")

HEADERS = ["Key", "Building Name", "Building Code", "Latitude", "Longitude", "Campus", "Source / Notes"]

# Key | Building Name | Code | Lat | Lon | Campus | Notes
ROWS = [
    ["D17", "Charles Perkins Centre", "D17", -33.887421, 151.183511, "On-campus (Camperdown)", "OSM"],
    ["J03", "Engineering & Technology Precinct (PNR Building)", "J03", -33.890141, 151.193116, "On-campus (Darlington)", "OSM (PNR) — verify exact building"],
    ["F09", "Madsen Building", "F09", -33.888715, 151.189449, "On-campus (Camperdown)", "OSM"],
    ["A31", "Sydney Nanoscience Hub", "A31", -33.888425, 151.187433, "On-campus (Camperdown)", "OSM"],
    ["G08", "Molecular Bioscience Building", "G08", -33.890260, 151.191191, "On-campus (Camperdown)", "OSM (Biochem & Microbiology) — verify"],
    ["F11", "Chemistry Building / School of Chemistry", "F11", -33.887901, 151.189550, "On-campus (Camperdown)", "OSM"],
    ["A28", "Physics Building", "A28", -33.888163, 151.187129, "On-campus (Camperdown)", "OSM"],
    ["A10ma", "Macleay Building (A10)", "A10ma", -33.885113, 151.188382, "On-campus (Camperdown)", "VERIFY — A10 building uncertain"],
    ["J07", "School of Aerospace, Mechanical & Mechatronic Eng.", "J07", -33.889300, 151.192500, "On-campus (Darlington)", "Estimate — verify"],
    ["J05", "PNR / Link area (Workshop)", "J05", -33.890141, 151.193116, "On-campus (Darlington)", "Estimate (near PNR) — verify"],
    ["D18", "Susan Wakil Health Building", "D18", -33.889083, 151.184661, "On-campus (Camperdown)", "OSM"],
    ["Centenary Institute", "Centenary Institute", "", -33.888086, 151.183886, "Off-campus (RPA, Camperdown)", "OSM"],
    ["Brain and Mind Centre", "Brain and Mind Centre (94 Mallett St)", "", -33.886831, 151.176570, "Off-campus (Camperdown)", "Street-level — verify"],
    ["Kolling Institute", "Kolling Institute", "", -33.820579, 151.191015, "Off-campus (St Leonards)", "OSM"],
    ["Royal North Shore Hospital", "Royal North Shore Hospital", "", -33.821432, 151.191120, "Off-campus (St Leonards)", "OSM"],
    ["Sydney Institute of Agriculture", "Sydney Institute of Agriculture (distributed)", "", -33.888500, 151.187500, "Off-campus / distributed", "Placeholder at campus — verify"],
    ["Biomedical Building, Australian Technology Park", "Biomedical Building, ATP", "", -33.896263, 151.195969, "Off-campus (South Eveleigh)", "OSM"],
    ["Moore College", "Moore Theological College", "", -33.891704, 151.187057, "Off-campus (Newtown)", "OSM"],
    ["Narrabri Campus", "I.A. Watson Grains Research Centre", "", -30.349594, 149.756760, "Regional (Narrabri)", "Approx (Newell Hwy) — verify"],
    # Geoscience (Map Data) has no building/location in source — left blank to fill in.
    ["Geoscience (TSS) — NO LOCATION", "Geoscience", "", None, None, "Unknown", "Map Data row has no building; add lat/lon to map it"],
]

wb = load_workbook(WB)
if "Buildings" in wb.sheetnames:
    del wb["Buildings"]
ws = wb.create_sheet("Buildings")

head_font = Font(bold=True, color="FFFFFF")
head_fill = PatternFill("solid", fgColor="424242")  # corporate charcoal
ws.append(HEADERS)
for c in range(1, len(HEADERS) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(vertical="center")

verify_fill = PatternFill("solid", fgColor="FFF3CD")  # soft amber for rows to verify
for r in ROWS:
    ws.append(r)
    row_idx = ws.max_row
    note = str(r[6]).lower()
    if "verify" in note or "no location" in note or "placeholder" in note:
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=c).fill = verify_fill

# Number format + widths + freeze header
for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
    for cell in row:
        cell.number_format = "0.000000"
widths = [34, 46, 14, 13, 13, 30, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# Put Buildings right after Map Data for discoverability
order = wb.sheetnames
if "Map Data" in order:
    wb.move_sheet("Buildings", offset=-(len(order) - 1 - order.index("Map Data")))

wb.save(WB)
print(f"Buildings sheet written: {len(ROWS)} rows. Sheets now: {wb.sheetnames}")
