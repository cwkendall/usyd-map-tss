"""Fill in missing building coordinates in the workbook's 'Buildings' sheet.

Idempotent maintenance step: when new facilities/locations are added to 'Map Data'
(a new Building Code, or a new off-campus Location/Building name), run this to
geocode just the missing ones and append them to the 'Buildings' sheet, flagged
amber for verification. It NEVER overwrites rows that already have coordinates, so
your hand-verified lat/lon are safe.

  - Reads 'Map Data' to find every location key (Building Code, else Location name),
    using the same join key as scripts/build-data.mjs.
  - Compares against existing 'Buildings' rows that already have lat/lon.
  - Geocodes only the missing keys via OpenStreetMap Nominatim (curated queries for
    known USyd buildings; name+address fallback otherwise). Respectful: sequential,
    1.1s delay, descriptive User-Agent.
  - Writes results back with openpyxl (preserves the styles of the other sheets).

Run sandboxed (per repo policy):
    uv run --with openpyxl python scripts/geocode-buildings.py

Network is only touched for keys that are actually missing, so re-running when
nothing new was added makes zero requests.
"""
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

WB = os.path.join(os.path.dirname(__file__), "..", "TSS-CRF-MapData.xlsx")
UA = "usyd-tss-crf-map/0.1 (geocode-buildings; renee.e.barber@gmail.com)"

# Known-good queries for USyd buildings/sites that Nominatim's free-text search
# resolves poorly. New entries only need adding here if the generic fallback fails.
CURATED = {
    "D17": "Charles Perkins Centre, University of Sydney, Camperdown NSW",
    "J03": "PNR Building, University of Sydney, Darlington NSW",
    "F09": "Madsen Building, University of Sydney, Camperdown NSW",
    "A31": "Sydney Nanoscience Hub, University of Sydney, Camperdown NSW",
    "G08": "Molecular Bioscience Building, University of Sydney, Camperdown NSW",
    "F11": "School of Chemistry, University of Sydney, Camperdown NSW",
    "A28": "Physics Building, University of Sydney, Camperdown NSW",
    "A10ma": "Macleay Building, University of Sydney, Camperdown NSW",
    "J07": "School of Aerospace Mechanical and Mechatronic Engineering, University of Sydney",
    "J05": "Link Building, University of Sydney, Darlington NSW",
    "D18": "Susan Wakil Health Building, University of Sydney, Camperdown NSW",
    "Centenary Institute": "Centenary Institute, Missenden Road, Camperdown NSW",
    "Brain and Mind Centre": "94 Mallett Street, Camperdown NSW 2050",
    "Kolling Institute": "Kolling Building, St Leonards NSW",
    "Royal North Shore Hospital": "Royal North Shore Hospital, St Leonards NSW",
    "Biomedical Building, Australian Technology Park": "Biomedical Building, Australian Technology Park, Eveleigh NSW",
    "Moore College": "Moore Theological College, Newtown NSW",
    "Narrabri Campus": "I A Watson Grains Research Centre, Narrabri NSW",
    "Sydney Institute of Agriculture": "RMC Gunn Building, University of Sydney, Camperdown NSW",
}


def cell(v):
    return "" if v is None else str(v).strip()


def geocode(query):
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"format": "jsonv2", "limit": 1, "countrycodes": "au", "q": query}
    )
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=20) as r:
        data = json.load(r)
    if not data:
        return None
    return float(data[0]["lat"]), float(data[0]["lon"]), data[0]["display_name"]


def queries_for(key, name, address, on_campus):
    if key in CURATED:
        return [CURATED[key]]
    out = []
    if name:
        if on_campus.lower().startswith("y"):
            out.append(f"{name}, University of Sydney, NSW Australia")
        if address:
            chunk = address.split("—")[0].split(";")[0].strip()
            if chunk:
                out.append(f"{name}, {chunk}, NSW Australia")
                out.append(f"{chunk}, NSW Australia")
        out.append(f"{name}, Sydney NSW Australia")
    # de-dup, keep order
    seen, uniq = set(), []
    for q in out:
        if q not in seen:
            seen.add(q)
            uniq.append(q)
    return uniq


def main():
    wb = load_workbook(WB)
    if "Map Data" not in wb.sheetnames or "Buildings" not in wb.sheetnames:
        sys.exit("Workbook must have 'Map Data' and 'Buildings' sheets.")
    md = wb["Map Data"]
    bs = wb["Buildings"]

    def header_index(ws):
        return {cell(c.value): i for i, c in enumerate(ws[1], start=1) if cell(c.value)}

    mi = header_index(md)
    bi = header_index(bs)

    # Existing Buildings keys that already have BOTH lat and lon.
    have_coords = set()
    all_keys_in_bs = set()
    for row in bs.iter_rows(min_row=2):
        k = cell(row[bi["Key"] - 1].value)
        if not k:
            continue
        all_keys_in_bs.add(k)
        lat = row[bi["Latitude"] - 1].value
        lon = row[bi["Longitude"] - 1].value
        if lat not in (None, "") and lon not in (None, ""):
            have_coords.add(k)

    # Unique location keys from Map Data (same join key as build-data.mjs),
    # capturing a representative name/address/on-campus for query building.
    needed = {}  # key -> dict(name, code, address, on_campus)
    for row in md.iter_rows(min_row=2):
        code = cell(row[mi["Building Code"] - 1].value)
        loc = cell(row[mi["Location / Building"] - 1].value)
        key = code or loc
        if not key:
            continue
        needed.setdefault(
            key,
            {
                "name": loc or code,
                "code": code,
                "address": cell(row[mi["Address / Notes"] - 1].value),
                "on_campus": cell(row[mi["On Campus"] - 1].value),
            },
        )

    missing = [k for k in needed if k not in have_coords]
    if not missing:
        print(f"✔ All {len(needed)} Map Data locations already have coordinates. No geocoding needed.")
        return

    print(f"Geocoding {len(missing)} location(s) missing coordinates…\n")
    verify_fill = PatternFill("solid", fgColor="FFF3CD")
    headers = [h for h, _ in sorted(bi.items(), key=lambda x: x[1])]
    added, failed = 0, 0

    for key in missing:
        info = needed[key]
        result = None
        for q in queries_for(key, info["name"], info["address"], info["on_campus"]):
            try:
                result = geocode(q)
            except Exception as e:  # noqa: BLE001
                print(f"  ! {key}: request error ({e})")
            if result:
                break
            time.sleep(1.1)
        campus = "On-campus" if info["on_campus"].lower().startswith("y") else "Off-campus"
        if result:
            lat, lon, disp = result
            note = f"Auto-geocoded {date.today()} via OSM — VERIFY"
            added += 1
            print(f"  ✔ {key:<32} {lat:.6f}, {lon:.6f}  {disp[:55]}")
        else:
            lat = lon = None
            note = "NOT FOUND — add lat/lon manually"
            failed += 1
            print(f"  ✘ {key:<32} NOT FOUND — fill manually")

        values = {
            "Key": key,
            "Building Name": info["name"],
            "Building Code": info["code"],
            "Latitude": lat,
            "Longitude": lon,
            "Campus": campus,
            "Source / Notes": note,
        }
        new_row = [values.get(h, "") for h in headers]
        bs.append(new_row)
        r = bs.max_row
        for c in range(1, len(headers) + 1):
            bs.cell(row=r, column=c).fill = verify_fill
        if "Latitude" in bi:
            bs.cell(row=r, column=bi["Latitude"]).number_format = "0.000000"
            bs.cell(row=r, column=bi["Longitude"]).number_format = "0.000000"
        time.sleep(1.1)

    wb.save(WB)
    print(f"\nDone. Added {added} geocoded, {failed} not-found. Review the amber rows in 'Buildings', then run `npm run build:data`.")


if __name__ == "__main__":
    main()
